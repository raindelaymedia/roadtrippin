"""
Road Trippin' — Revenue Tracker (CSV)  ·  one-stop monthly revenue builder
Interactive prompt that appends one row per source to data/revenue.csv,
PLUS automatic ingestion of the monthly Backyard Ventures payout sheet.
Use multiple times to backfill historical months.

CSV format:
    period,source,amount
    2026-03,YT_VIDEOS,3528.49
    2026-03,YT_SHORTS,284.47
    2026-03,CULTURE_GENESIS,TBD

Usage:
    python build_revenue_tracker.py
    python build_revenue_tracker.py --csv data/revenue.csv
    python build_revenue_tracker.py --backyard data/BackyardRevMay.xlsx
    python build_revenue_tracker.py --backyard data/BackyardRevJun.xlsx --skip-prompt

Backyard Ventures:
    The script will look for a Backyard sheet (via --backyard, or it asks
    interactively). It keeps only rows with a populated "Date Paid", converts
    the Excel serial date, buckets by paid month, splits Podcast vs
    Programmatic, and writes BACKYARD_PODCAST / BACKYARD_PROGRAMMATIC rows.
    Ingestion is idempotent per month, so re-running with an updated sheet is
    safe. These two sources are managed only by the ingest — they are never
    prompted, and the manual flow never overwrites them.

Requirements:
    pip install openpyxl
"""

import argparse
import csv
import os
from collections import defaultdict
from datetime import datetime, date, timedelta


# Sources in order — matches the original Excel tracker layout
SOURCES = [
    "YT_VIDEOS",
    "YT_SHORTS",
    "YT_LIVES",
    "CULTURE_GENESIS",
    "UPTIDES",
    "ACAST",
    "FANATICS",
    "SOCIAL",
    "MERCH",
]

# Backyard Ventures — auto-ingested from xlsx, never prompted.
BACKYARD_SOURCES = ["BACKYARD_PODCAST", "BACKYARD_PROGRAMMATIC"]
# Revenue Type value (lowercased) -> revenue.csv source key
BACKYARD_TYPE_TO_SOURCE = {
    "podcast":      "BACKYARD_PODCAST",
    "programmatic": "BACKYARD_PROGRAMMATIC",
}
# Default sheet to look for if --backyard isn't supplied.
BACKYARD_DEFAULT_FILE = "BackyardRevMay.xlsx"
# Excel's day-zero is 1899-12-30 (the 1900 leap-year bug baked in).
EXCEL_EPOCH = date(1899, 12, 30)

# 3-letter month codes for friendlier prompts
MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
          "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def parse_value(raw):
    """Parse user input into a CSV-friendly string."""
    s = raw.strip()
    if not s:
        return ""  # blank → leave empty
    s_lower = s.lower()
    if s_lower in ("na", "n/a"):
        return "N/A"
    if s_lower == "tbd":
        return "TBD"
    try:
        return f"{float(s.replace(',', '').replace('$', '')):.2f}"
    except ValueError:
        return s  # whatever the user typed


def prompt_period():
    today = datetime.now()
    default_month = today.strftime("%b").upper()
    default_year = today.year

    month_in = input(f"Month [{default_month}]: ").strip().upper() or default_month
    year_in = input(f"Year [{default_year}]: ").strip()
    year = int(year_in) if year_in else default_year

    if month_in not in MONTHS:
        # Try to be forgiving — match prefix
        match = next((m for m in MONTHS if m.startswith(month_in[:3])), None)
        if match:
            month_in = match
        else:
            raise ValueError(f"Unknown month: {month_in}")

    month_num = MONTHS.index(month_in) + 1
    period = f"{year}-{month_num:02d}"
    return period


# ─── Backyard Ventures ingestion ─────────────────────────────────
def _excel_serial_to_month(value):
    """Convert an Excel date (serial int/float or real datetime) to 'YYYY-MM'.
    Returns None if the value isn't a usable date (e.g. unpaid row)."""
    if value is None or value == "":
        return None
    if hasattr(value, "year"):  # openpyxl gave us a real datetime
        return f"{value.year:04d}-{value.month:02d}"
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    if serial <= 0:
        return None
    d = EXCEL_EPOCH + timedelta(days=int(serial))
    return f"{d.year:04d}-{d.month:02d}"


def _find_backyard_columns(header_row):
    """Map needed Backyard columns to 0-based indices. Raises if missing."""
    norm = [(str(h).strip().lower() if h is not None else "") for h in header_row]
    wanted = {"rev_type": "revenue type", "due": "due to host", "paid": "date paid"}
    idx = {}
    for key, target in wanted.items():
        try:
            idx[key] = norm.index(target)
        except ValueError:
            raise SystemExit(
                f"\n⚠  Backyard sheet is missing a '{target}' column.\n"
                f"   Headers found: {[h for h in header_row if h]}\n"
            )
    return idx


def read_backyard(path):
    """Return ({(month, source): total_due}, counted, skipped_unpaid, skipped_other)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("\n⚠  openpyxl not installed.  Run: pip install openpyxl\n")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise SystemExit(f"\n⚠  {path} appears to be empty.\n")

    idx = _find_backyard_columns(header)
    totals = defaultdict(float)
    counted = skipped_unpaid = skipped_other = 0

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        month = _excel_serial_to_month(row[idx["paid"]])
        if month is None:                       # not yet paid -> skip
            skipped_unpaid += 1
            continue
        rev_type = row[idx["rev_type"]]
        type_key = str(rev_type).strip().lower() if rev_type is not None else ""
        source = BACKYARD_TYPE_TO_SOURCE.get(type_key)
        if source is None:                       # unrecognized revenue type
            skipped_other += 1
            continue
        try:
            amount = float(row[idx["due"]])
        except (TypeError, ValueError):
            skipped_other += 1
            continue
        totals[(month, source)] += amount
        counted += 1

    wb.close()
    return totals, counted, skipped_unpaid, skipped_other


def ingest_backyard(existing, path):
    """Fold a Backyard sheet into the `existing` {(period,source):amount} dict.
    Idempotent per month: replaces BACKYARD_* rows for months the sheet covers.
    Returns the list of months touched (for reporting)."""
    totals, counted, skipped_unpaid, skipped_other = read_backyard(path)
    if not totals:
        print(f"  No paid Backyard rows found in {os.path.basename(path)} "
              f"({skipped_unpaid} unpaid skipped).")
        return []

    months_touched = sorted({m for (m, _) in totals})

    # Drop existing BACKYARD_* rows for the covered months, then add fresh ones.
    for (p, s) in list(existing.keys()):
        if s in BACKYARD_SOURCES and p in months_touched:
            del existing[(p, s)]
    for (month, source), amount in totals.items():
        existing[(month, source)] = f"{amount:.2f}"

    print(f"  Paid rows: {counted}  ·  unpaid skipped: {skipped_unpaid}  ·  "
          f"other skipped: {skipped_other}")
    for m in months_touched:
        pod = totals.get((m, "BACKYARD_PODCAST"), 0.0)
        prg = totals.get((m, "BACKYARD_PROGRAMMATIC"), 0.0)
        print(f"    {m}   Podcast ${pod:>9,.2f}   Programmatic ${prg:>9,.2f}   "
              f"Total ${pod + prg:>9,.2f}")
    return months_touched


def load_existing(csv_path):
    """Load existing CSV into {(period, source): amount} dict."""
    if not os.path.exists(csv_path):
        return {}
    rows = {}
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows[(r["period"], r["source"])] = r["amount"]
    return rows


def write_csv(csv_path, data):
    """Write all rows back to CSV, sorted newest-first."""
    rows = sorted(data.items(), key=lambda kv: (kv[0][0], SOURCES.index(kv[0][1]) if kv[0][1] in SOURCES else 99), reverse=False)
    # Sort by period DESC, then by source order
    rows.sort(key=lambda kv: (-int(kv[0][0].replace("-", "")),
                               SOURCES.index(kv[0][1]) if kv[0][1] in SOURCES else 99))
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["period", "source", "amount"])
        for (period, source), amount in rows:
            writer.writerow([period, source, amount])


def prompt_sources(period, existing):
    print(f"\nEntering revenue for {period}")
    print("(blank = skip, 'tbd' = TBD, 'na' = N/A)\n")

    new_data = {}

    # Standard sources first
    for source in SOURCES:
        existing_val = existing.get((period, source), "")
        if existing_val:
            prompt = f"  {source:<18} [{existing_val}] "
        else:
            prompt = f"  {source:<18} "
        raw = input(prompt)
        if not raw.strip() and existing_val:
            new_data[source] = existing_val
        else:
            val = parse_value(raw)
            if val:
                new_data[source] = val

    # Surface any custom sources that already have data for this period
    # (Backyard rows are ingest-managed, so never surface them for manual edit).
    period_custom = {s: a for (p, s), a in existing.items()
                     if p == period and s not in SOURCES and s not in BACKYARD_SOURCES}
    if period_custom:
        print("\nExisting custom sources for this period:")
        for source in sorted(period_custom.keys()):
            existing_val = period_custom[source]
            raw = input(f"  {source:<18} [{existing_val}] ")
            if not raw.strip():
                new_data[source] = existing_val
            else:
                val = parse_value(raw)
                if val:
                    new_data[source] = val

    # Prompt for new custom sources
    new_data.update(prompt_custom_sources(existing))

    return new_data


def prompt_custom_sources(existing):
    """Loop prompt for adding new custom revenue sources."""
    custom_data = {}
    # Show what custom sources have ever been used (for reference)
    all_custom = sorted(set(s for (_, s) in existing.keys() if s not in SOURCES))

    print("")
    while True:
        ans = input("Add a custom source? [y/N]: ").strip().lower()
        if ans not in ("y", "yes"):
            break

        if all_custom:
            print(f"  (Previously used: {', '.join(all_custom)})")

        name_raw = input("  Source name (e.g. PATREON, LIVE_EVENTS): ").strip().upper()
        if not name_raw:
            print("  Skipped — empty name.")
            continue

        # Sanitize: replace spaces, slashes, and dashes with underscores
        name = name_raw.replace(" ", "_").replace("-", "_").replace("/", "_")
        # Collapse repeated underscores
        while "__" in name:
            name = name.replace("__", "_")

        if name in SOURCES:
            print(f"  '{name}' is already a standard source. Use the prompt above instead.")
            continue

        amount_raw = input(f"  {name} amount: ").strip()
        val = parse_value(amount_raw)
        if val:
            custom_data[name] = val
            print(f"  ✓ Added {name} = {val}")
        else:
            print(f"  Skipped — empty/invalid amount.")

    return custom_data


def calculate_total(period_data):
    """Sum all numeric values in the period."""
    total = 0.0
    for v in period_data.values():
        try:
            total += float(v)
        except (ValueError, TypeError):
            pass
    return total


def main():
    parser = argparse.ArgumentParser(description="Road Trippin' Revenue Tracker (CSV)")
    parser.add_argument("--csv", default=None, help="Path to revenue.csv")
    parser.add_argument("--backyard", default=None,
                        help="Path to a Backyard Ventures .xlsx to ingest")
    parser.add_argument("--skip-prompt", action="store_true",
                        help="Only ingest Backyard (skip the interactive per-source prompts)")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    csv_path = args.csv or os.path.join(data_dir, "revenue.csv")

    print("=" * 55)
    print("ROAD TRIPPIN' — REVENUE TRACKER (CSV)")
    print("=" * 55)
    print(f"CSV: {csv_path}\n")

    existing = load_existing(csv_path)
    if existing:
        periods = sorted(set(p for (p, _) in existing.keys()), reverse=True)
        print(f"Existing data: {len(existing)} rows across {len(periods)} months")
        print(f"Most recent: {periods[0]}\n")

    # ─── Backyard Ventures ingest ──────────────────────────────────
    # Resolve a sheet path: --backyard wins; otherwise offer the default if it
    # exists, or let the user type a path. Blank input skips ingestion.
    backyard_path = args.backyard
    if backyard_path is None:
        default_path = os.path.join(data_dir, BACKYARD_DEFAULT_FILE)
        suggestion = BACKYARD_DEFAULT_FILE if os.path.exists(default_path) else ""
        prompt = (f"Backyard Ventures sheet to ingest [{suggestion}] "
                  if suggestion else "Backyard Ventures sheet to ingest (blank to skip): ")
        raw = input(prompt).strip()
        if raw:
            backyard_path = raw
        elif suggestion:
            backyard_path = default_path

    if backyard_path:
        if not os.path.isabs(backyard_path):
            # try as given, then relative to data_dir
            cand = backyard_path if os.path.exists(backyard_path) \
                else os.path.join(data_dir, backyard_path)
            backyard_path = cand
        if os.path.exists(backyard_path):
            print(f"\n─── Ingesting Backyard Ventures: {os.path.basename(backyard_path)} ───")
            ingest_backyard(existing, backyard_path)
        else:
            print(f"\n⚠ Backyard sheet not found: {backyard_path} — skipping ingest.")
        # Persist the ingest immediately so it survives even if the user
        # cancels the manual prompts below.
        write_csv(csv_path, existing)

    if args.skip_prompt:
        print(f"\n✓ Backyard ingest saved to {csv_path}")
        print("=" * 55)
        return

    # Prompt for period
    period = prompt_period()

    # Check if exists
    period_existing = {s: a for (p, s), a in existing.items() if p == period}
    if period_existing:
        print(f"\n⚠ {period} already has {len(period_existing)} entries.")
        print("  Press enter on any source to keep existing value.\n")

    # Prompt for sources
    new_data = prompt_sources(period, existing)

    # Preview
    total = calculate_total(new_data)
    print(f"\n─── Preview for {period} ───")
    for source in SOURCES:
        val = new_data.get(source, "—")
        print(f"  {source:<18} {val}")
    # Show any custom sources separately
    custom_in_data = {s: v for s, v in new_data.items() if s not in SOURCES}
    if custom_in_data:
        print("  ─── Custom sources ───")
        for source, val in sorted(custom_in_data.items()):
            print(f"  {source:<18} {val}")
    print(f"  {'TOTAL':<18} ${total:,.2f}")

    # Confirm
    confirm = input("\nSave? [y/N]: ").strip().lower()
    if confirm not in ("y", "yes"):
        print("Cancelled.")
        return

    # Merge new data into existing — replace any prior values for this period.
    # BACKYARD_* rows are owned by the ingest, so never delete them here.
    for (p, s) in list(existing.keys()):
        if p == period and s not in BACKYARD_SOURCES:
            del existing[(p, s)]
    for source, amount in new_data.items():
        existing[(period, source)] = amount

    write_csv(csv_path, existing)
    print(f"\n✓ Saved {len(new_data)} rows for {period}")
    print(f"  CSV: {csv_path}")
    print("=" * 55)


if __name__ == "__main__":
    main()