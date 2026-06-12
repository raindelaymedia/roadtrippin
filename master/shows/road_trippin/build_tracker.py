"""
Road Trippin' — Analytics Tracker Builder v2
Uses YouTube Analytics API (via OAuth Playground refresh token) + Megaphone CSVs
to generate an Excel tracker matching the original format exactly.

Usage:
    python build_tracker.py
    python build_tracker.py --months 22
    python build_tracker.py --megaphone-monthly data/megaphone01-042026.csv

Requirements:
    pip install google-auth google-auth-oauthlib google-api-python-client openpyxl python-dateutil
"""

import argparse
import csv
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build as build_api

# ─── Credentials (loaded from config.py — never hardcoded) ──────
try:
    from config import (
        YT_CHANNEL_ID    as CHANNEL_ID,
        YT_CLIENT_ID     as CLIENT_ID,
        YT_CLIENT_SECRET as CLIENT_SECRET,
        YT_REFRESH_TOKEN as REFRESH_TOKEN,
        MEGAPHONE_TOKEN,
        MEGAPHONE_NETWORK_ID,
        MEGAPHONE_PODCAST_ID,
        MEGAPHONE_API_BASE,
    )
except ImportError:
    raise SystemExit(
        "\n⚠  config.py not found.\n"
        "   Copy config.example.py → config.py and fill in your credentials.\n"
    )

# ─── Styling (matches original tracker) ──────────────────────────
FONT_TITLE = Font(name="Barlow Condensed", size=18, bold=True)
FONT_SECTION_LABEL = Font(name="Barlow Condensed", size=13)
FONT_MONTH = Font(name="Barlow Condensed", size=13)
FONT_ROW_LABEL = Font(name="Barlow Condensed", size=12, bold=True)
FONT_SUB_LABEL = Font(name="Barlow Condensed", size=12, bold=False)
FONT_DATA = Font(name="Barlow Condensed", size=13)
FONT_NA = Font(name="Barlow Condensed", size=13, color="D9D9D9")

FILL_ALT = PatternFill("solid", fgColor="F3F3F3")
FILL_SECTION = PatternFill("solid", fgColor="D9D9D9")
FILL_NONE = PatternFill(fill_type=None)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'


# ─── Month helpers ───────────────────────────────────────────────
def get_month_keys(n=22):
    now = datetime.now().replace(day=1)
    return [(now - relativedelta(months=i)).strftime("%Y-%m") for i in range(n)]


def month_key_to_label(ym):
    dt = datetime.strptime(ym, "%Y-%m")
    now = datetime.now()
    label = dt.strftime("%b").upper()
    if dt.year < now.year - 1:
        label += f" {dt.strftime('%y')}"
    elif dt.year < now.year:
        label += f" {dt.strftime('%y')}"
    return label


# ─── YouTube Analytics API ───────────────────────────────────────
def authenticate_yt():
    creds = Credentials(
        token=None, refresh_token=REFRESH_TOKEN,
        client_id=CLIENT_ID, client_secret=CLIENT_SECRET,
        token_uri="https://oauth2.googleapis.com/token",
        scopes=["https://www.googleapis.com/auth/yt-analytics.readonly",
                "https://www.googleapis.com/auth/youtube.readonly"])
    creds.refresh(Request())
    return creds


def pull_yt_monthly(yt_analytics, start_date, end_date):
    resp = yt_analytics.reports().query(
        ids="channel==MINE", startDate=start_date, endDate=end_date,
        metrics="views,estimatedMinutesWatched,subscribersGained,subscribersLost",
        dimensions="month", sort="month").execute()
    data = {}
    for row in resp.get("rows", []):
        data[row[0]] = {"views": row[1], "watch_hrs": round(row[2] / 60, 1),
                        "subs_gained": row[3], "subs_lost": row[4]}
    return data


def pull_yt_by_content_type(yt_analytics, start_date, end_date):
    resp = yt_analytics.reports().query(
        ids="channel==MINE", startDate=start_date, endDate=end_date,
        metrics="views,estimatedMinutesWatched",
        dimensions="month,creatorContentType", sort="month").execute()
    data = defaultdict(lambda: {"VIDEO_ON_DEMAND": 0, "SHORTS": 0, "LIVE_STREAM": 0})
    rows = resp.get("rows", [])
    ctypes = set(row[1] for row in rows)
    print(f"    Content types from API: {ctypes}")
    # API returns camelCase: videoOnDemand, shorts, liveStream, posts
    KEY_MAP = {
        "videoOnDemand": "VIDEO_ON_DEMAND",
        "shorts":        "SHORTS",
        "liveStream":    "LIVE_STREAM",
        "posts":         "POSTS",
    }
    for row in rows:
        key = KEY_MAP.get(row[1], row[1])
        data[row[0]][key] = data[row[0]].get(key, 0) + row[2]
    return dict(data)


def pull_yt_subscriber_total(youtube):
    resp = youtube.channels().list(part="statistics", id=CHANNEL_ID).execute()
    return int(resp["items"][0]["statistics"].get("subscriberCount", 0))


def pull_yt_shorts_count(youtube):
    """Count Shorts published per month using the Data API uploads playlist."""
    import re as _re
    ch = youtube.channels().list(part="contentDetails", id=CHANNEL_ID).execute()
    uploads_id = ch["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

    counts = defaultdict(int)
    next_page = None
    while True:
        resp = youtube.playlistItems().list(
            part="contentDetails", playlistId=uploads_id,
            maxResults=50, pageToken=next_page).execute()
        video_ids = [item["contentDetails"]["videoId"] for item in resp["items"]]

        # Get durations to identify Shorts (<= 60s officially, but up to ~3min tagged as shorts)
        details = youtube.videos().list(
            part="contentDetails,snippet", id=",".join(video_ids)).execute()
        for vid in details["items"]:
            dur = vid.get("contentDetails", {}).get("duration", "") or ""
            if not dur: continue   # active live, premiere, or post — skip
            # Parse ISO 8601 duration
            m = _re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", dur)
            seconds = int(m.group(1) or 0)*3600 + int(m.group(2) or 0)*60 + int(m.group(3) or 0) if m else 0
            # Shorts: <= 180 seconds (3 min) — matches our content type classifier
            if seconds <= 180:
                pub_month = vid["snippet"]["publishedAt"][:7]
                counts[pub_month] += 1

        next_page = resp.get("nextPageToken")
        if not next_page:
            break

    return dict(counts)


def pull_top_content_monthly(youtube, months, top_n=10):
    """Single full walk of the uploads playlist, then bucket every video by its
    publish month and content type (short/mid/long), and keep the top N per
    bucket per month.

    Same 3/15-minute boundaries as pull_best_of so classification is consistent
    across the dashboard.  `months` is the tracker's list of YYYY-MM keys —
    the oldest is used as a pagination cutoff so we don't walk before the
    tracker window.

    Returns: { "YYYY-MM": { "long":[...], "mid":[...], "short":[...] }, ... }
    """
    import re as _re

    oldest = min(months) if months else "1970-01"
    cutoff = f"{oldest}-01T00:00:00Z"

    ch = youtube.channels().list(part="contentDetails", id=CHANNEL_ID).execute()
    uploads_id = ch["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

    videos, seen_ids, next_page = [], set(), None
    while True:
        resp = youtube.playlistItems().list(
            part="contentDetails,snippet", playlistId=uploads_id,
            maxResults=50, pageToken=next_page).execute()
        video_ids = [it["contentDetails"]["videoId"] for it in resp["items"]]
        if not video_ids:
            break
        details = youtube.videos().list(
            part="contentDetails,snippet,statistics", id=",".join(video_ids)).execute()

        stop = False
        for vid in details["items"]:
            vid_id = vid["id"]
            if vid_id in seen_ids:
                continue
            seen_ids.add(vid_id)
            pub = vid["snippet"]["publishedAt"]
            if pub < cutoff:
                stop = True
                continue
            dur = vid.get("contentDetails", {}).get("duration", "") or ""
            if not dur: continue   # active live, premiere, or post — skip
            m = _re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", dur)
            secs = int(m.group(1) or 0)*3600 + int(m.group(2) or 0)*60 + int(m.group(3) or 0) if m else 0
            if secs <= 180:
                vtype = "short"
            elif secs <= 1800:
                vtype = "mid"
            else:
                vtype = "long"
            videos.append({
                "id":        vid_id,
                "title":     vid["snippet"]["title"],
                "published": pub[:10],
                "month":     pub[:7],
                "views":     int(vid["statistics"].get("viewCount", 0)),
                "likes":     int(vid["statistics"].get("likeCount", 0)),
                "thumbnail": (
                    vid["snippet"].get("thumbnails", {}).get("maxres", {}).get("url") or
                    vid["snippet"].get("thumbnails", {}).get("high",   {}).get("url") or
                    f"https://img.youtube.com/vi/{vid_id}/hqdefault.jpg"
                ),
                "url":  f"https://youtu.be/{vid_id}",
                "type": vtype,
            })
        next_page = resp.get("nextPageToken")
        if stop or not next_page:
            break

    monthly = {}
    for v in videos:
        m = v["month"]
        monthly.setdefault(m, {"long": [], "mid": [], "short": []})
        monthly[m][v["type"]].append(v)
    for m in monthly:
        for t in ("long", "mid", "short"):
            monthly[m][t] = sorted(monthly[m][t], key=lambda x: x["views"], reverse=True)[:top_n]
    return monthly


def pull_best_of(youtube, months, n_months=1):
    """Pull top performing videos per content type for the most recent n_months.
    Returns dict with keys: long, mid, short — each a list of video dicts."""
    import re as _re
    from datetime import datetime as _dt

    # Calculate date range for most recent n_months
    now = _dt.now()
    cutoff = (now.replace(day=1) - relativedelta(months=n_months - 1)).strftime("%Y-%m-%dT00:00:00Z")

    ch = youtube.channels().list(part="contentDetails", id=CHANNEL_ID).execute()
    uploads_id = ch["items"][0]["contentDetails"]["relatedPlaylists"]["uploads"]

    videos = []
    seen_ids = set()
    next_page = None
    while True:
        resp = youtube.playlistItems().list(
            part="contentDetails,snippet", playlistId=uploads_id,
            maxResults=50, pageToken=next_page).execute()

        video_ids = [item["contentDetails"]["videoId"] for item in resp["items"]]

        details = youtube.videos().list(
            part="contentDetails,snippet,statistics", id=",".join(video_ids)).execute()

        for vid in details["items"]:
            vid_id = vid["id"]
            if vid_id in seen_ids:
                continue
            seen_ids.add(vid_id)

            pub = vid["snippet"]["publishedAt"]
            if pub < cutoff:
                next_page = None
                break
            dur = vid.get("contentDetails", {}).get("duration", "") or ""
            if not dur: continue   # active live, premiere, or post — skip
            m = _re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", dur)
            secs = int(m.group(1) or 0)*3600 + int(m.group(2) or 0)*60 + int(m.group(3) or 0) if m else 0

            if secs <= 180:
                vtype = "SHORT"
            elif secs <= 1800:
                vtype = "MID"
            else:
                vtype = "LONG"

            videos.append({
                "id":        vid_id,
                "title":     vid["snippet"]["title"],
                "published": pub[:10],
                "views":     int(vid["statistics"].get("viewCount", 0)),
                "likes":     int(vid["statistics"].get("likeCount", 0)),
                "thumbnail": (
                    vid["snippet"].get("thumbnails", {}).get("maxres", {}).get("url") or
                    vid["snippet"].get("thumbnails", {}).get("high",   {}).get("url") or
                    f"https://img.youtube.com/vi/{vid_id}/hqdefault.jpg"
                ),
                "url":   f"https://youtu.be/{vid_id}",
                "type":  vtype,
                "show":  "GROUP",
            })

        next_page = resp.get("nextPageToken")
        if not next_page:
            break

    # Sort each type by views desc, take top 5
    long_vids  = sorted([v for v in videos if v["type"] == "LONG"],  key=lambda x: x["views"], reverse=True)[:5]
    mid_vids   = sorted([v for v in videos if v["type"] == "MID"],   key=lambda x: x["views"], reverse=True)[:5]
    short_vids = sorted([v for v in videos if v["type"] == "SHORT"],  key=lambda x: x["views"], reverse=True)[:5]

    return {"long": long_vids, "mid": mid_vids, "short": short_vids}


def build_best_of(wb, best_of_data, date_range_label):
    """Build the Best Of tab matching the original tracker format."""
    ws = wb.create_sheet("⭐ Best Of")
    ws.sheet_properties.tabColor = "FFD700"

    # Title
    ws.cell(row=1, column=1, value=f"⭐ BEST OF — {date_range_label}").font = FONT_TITLE
    ws.cell(row=1, column=1).fill = FILL_ALT
    ws.cell(row=1, column=1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    for col in range(1, 8):
        ws.cell(row=1, column=col).fill = FILL_ALT

    # Column headers
    headers = ["Rank", "Title", "Published", "Views", "Likes", "Type", "Show"]
    widths   = [6, 55, 12, 12, 10, 8, 8]
    for i, (h, w) in enumerate(zip(headers, widths)):
        c = ws.cell(row=3, column=i+1, value=h)
        c.font = FONT_ROW_LABEL
        c.fill = FILL_SECTION
        c.alignment = ALIGN_CENTER
        ws.column_dimensions[get_column_letter(i+1)].width = w

    sections = [
        ("★ TOP FULL EPISODES", best_of_data.get("long", [])),
        ("★ TOP SEGMENT CLIPS",  best_of_data.get("mid", [])),
        ("★ TOP SHORTS",         best_of_data.get("short", [])),
    ]

    current_row = 4
    for section_label, videos in sections:
        # Section header
        c = ws.cell(row=current_row, column=1, value=section_label)
        c.font = FONT_ROW_LABEL
        c.fill = FILL_SECTION
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=current_row, column=col).fill = FILL_SECTION
        current_row += 1

        if not videos:
            c = ws.cell(row=current_row, column=1, value="No data for this period")
            c.font = FONT_NA
            current_row += 2
            continue

        for rank, vid in enumerate(videos, 1):
            fill = FILL_ALT if rank % 2 == 0 else FILL_NONE
            data = [
                (f"#{rank}", ALIGN_CENTER),
                (vid["title"], Alignment(horizontal="left", vertical="center", wrap_text=True)),
                (vid["published"], ALIGN_CENTER),
                (vid["views"], ALIGN_CENTER),
                (vid["likes"], ALIGN_CENTER),
                (vid["type"], ALIGN_CENTER),
                (vid["show"], ALIGN_CENTER),
            ]
            for col, (val, align) in enumerate(data, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.font = FONT_DATA
                c.fill = fill
                c.alignment = align
                if col in (4, 5):
                    c.number_format = NUM_FMT
            ws.row_dimensions[current_row].height = 30
            current_row += 1

        current_row += 1  # spacer between sections

    # Data source note
    ws.cell(row=current_row + 1, column=1,
            value="DATA SOURCE: YouTube Data API — views reflect lifetime totals at time of export").font = FONT_NA

    return ws


# ─── YouTube Reporting API ───────────────────────────────────────
# Jobs created 2026-04-24 — data starts accumulating from this date forward
REPORTING_JOBS = {
    "reach":   "54612c77-5fa6-4d27-9c82-be1aaf9c67b9",  # channel_reach_combined_a1 (CTR, impressions)
    "traffic": "45b2e820-0bb4-4596-a744-63d173117042",  # channel_traffic_source_a3 (search % of views)
    "device":  "e56d8e8c-1730-4e8c-88d8-c9ff3de5f429",  # channel_device_os_a3 (TV views %)
}


def pull_reporting_api(creds, data_dir):
    """Download latest Reporting API CSVs for each job and parse into monthly data.
    Returns dict keyed by month (YYYY-MM) with CTR, impressions, search_pct, tv_pct."""
    import requests as _req
    import csv as _csv
    import io as _io
    from collections import defaultdict as _dd

    token = creds.token
    headers = {"Authorization": f"Bearer {token}"}
    monthly = _dd(lambda: {
        "impressions": 0, "ctr": None,
        "search_views": 0, "total_views": 0,
        "tv_views": 0,
    })

    for job_key, job_id in REPORTING_JOBS.items():
        # List available reports for this job
        url = f"https://youtubereporting.googleapis.com/v1/jobs/{job_id}/reports"
        resp = _req.get(url, headers=headers)
        if resp.status_code != 200:
            print(f"  ⚠ Reporting API {job_key}: {resp.status_code}")
            continue

        reports = resp.json().get("reports", [])
        if not reports:
            print(f"  ⚠ {job_key}: no reports yet (check back tomorrow)")
            continue

        # Use most recent report
        latest = sorted(reports, key=lambda r: r.get("endTime", ""))[-1]
        dl_url = latest["downloadUrl"]
        start = latest.get("startTime", "")[:7]
        end = latest.get("endTime", "")[:7]
        print(f"  ✓ {job_key}: report {start} → {end}")

        # Download and cache the CSV
        csv_path = os.path.join(data_dir, f"reporting_{job_key}.csv")
        dl = _req.get(dl_url, headers=headers)
        if dl.status_code != 200:
            print(f"  ⚠ {job_key}: download failed {dl.status_code}")
            continue

        with open(csv_path, "w", encoding="utf-8") as f:
            f.write(dl.text)

        # Parse CSV
        reader = _csv.DictReader(_io.StringIO(dl.text))
        for row in reader:
            # Date column is YYYYMMDD — convert to YYYY-MM
            date_col = row.get("date", row.get("day", ""))
            if not date_col or len(date_col) < 7:
                continue
            month = f"{date_col[:4]}-{date_col[4:6]}"

            if job_key == "reach":
                monthly[month]["impressions"] += int(row.get("impressions", 0) or 0)
                ctr = row.get("impressionClickThroughRate")
                if ctr:
                    # Average CTR across days (will be weighted avg)
                    existing = monthly[month]["ctr"]
                    monthly[month]["ctr"] = float(ctr) if existing is None else (existing + float(ctr)) / 2

            elif job_key == "traffic":
                views = int(row.get("views", 0) or 0)
                source = row.get("traffic_source_type", row.get("trafficSourceType", ""))
                monthly[month]["total_views"] += views
                if source in ("SUBSCRIBER", "YT_SEARCH", "4"):  # 4 = YouTube search
                    monthly[month]["search_views"] += views

            elif job_key == "device":
                views = int(row.get("views", 0) or 0)
                device = row.get("device_type", row.get("deviceType", ""))
                monthly[month]["total_views"] += views
                if device in ("TV", "game_console"):
                    monthly[month]["tv_views"] += views

    # Calculate percentages
    result = {}
    for month, d in monthly.items():
        search_pct = round(d["search_views"] / d["total_views"], 4) if d["total_views"] > 0 else None
        tv_pct = round(d["tv_views"] / d["total_views"], 4) if d["total_views"] > 0 else None
        result[month] = {
            "impressions": d["impressions"] or None,
            "ctr": round(d["ctr"], 4) if d["ctr"] else None,
            "search_pct": search_pct,
            "tv_pct": tv_pct,
        }

    return result


# ─── Megaphone ───────────────────────────────────────────────────
def pull_megaphone_api(token, network_id, podcast_id, n_months=22):
    """Pull episode counts per month from Megaphone API.
    Returns (monthly, episode_counts) — monthly is empty dict since
    download analytics aren't available via the standard API.
    episode_counts = {'2026-04': 13, ...}
    """
    import requests as _req
    from datetime import datetime as _dt

    headers = {"Authorization": f"Token {token}"}

    # Paginate through all episodes to count per month
    all_episodes = []
    page = 1
    while True:
        url = f"{MEGAPHONE_API_BASE}/networks/{network_id}/podcasts/{podcast_id}/episodes"
        resp = _req.get(url, headers=headers, params={"page": page, "per_page": 500})
        if resp.status_code != 200:
            raise RuntimeError(f"Megaphone API {resp.status_code}: {resp.text[:200]}")
        batch = resp.json()
        if not batch:
            break
        all_episodes.extend(batch)
        if len(batch) < 500:
            break
        page += 1
        if page > 20:
            break

    print(f"  ✓ Megaphone API: pulled {len(all_episodes)} total episodes")

    # Count episodes by publish month
    episode_counts = defaultdict(int)
    for ep in all_episodes:
        pub = ep.get("pubdate") or ep.get("createdAt")
        if not pub:
            continue
        try:
            dt = _dt.strptime(pub[:10], "%Y-%m-%d")
            episode_counts[dt.strftime("%Y-%m")] += 1
        except (ValueError, TypeError):
            continue

    return {}, dict(episode_counts)


def load_megaphone_episode_counts(path):
    """Count episodes published per month from per-episode CSV.
    Supports both PUBLISHED DATE format (April 20, 2026) and MONTH format (2026-04)."""
    if not path or not os.path.exists(path): return {}
    from datetime import datetime as _dt
    counts = defaultdict(int)
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            # Try PUBLISHED DATE column first (newer export format)
            pub = row.get("PUBLISHED DATE", "").strip()
            if pub:
                try:
                    m = _dt.strptime(pub, "%B %d, %Y").strftime("%Y-%m")
                except ValueError:
                    try:
                        m = _dt.strptime(pub, "%Y-%m-%d").strftime("%Y-%m")
                    except ValueError:
                        continue
            # Fall back to MONTH column (older export format)
            elif "MONTH" in row:
                m = row["MONTH"].strip()
            else:
                continue
            counts[m] += 1
    return dict(counts)


def load_megaphone_monthly(path):
    if not path or not os.path.exists(path): return {}
    data = {}
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            # Handle both old ("DOWNLOADS") and new ("TOTAL DOWNLOADS") column names
            dl_raw = row.get("DOWNLOADS") or row.get("TOTAL DOWNLOADS") or "0"
            st_raw = row.get("STREAMS") or row.get("TOTAL STREAMS") or "0"
            td_raw = row.get("TOTAL DELIVERY") or "0"
            dl = int(dl_raw.replace(",", "") or 0)
            st = int(st_raw.replace(",", "") or 0)
            td = int(td_raw.replace(",", "") or 0) or dl
            # Normalize month key to YYYY-MM regardless of input format
            month_raw = row.get("MONTH", "").strip()
            month_key = month_raw[:7] if len(month_raw) >= 7 else month_raw
            if month_key:
                data[month_key] = {"downloads": dl, "streams": st, "total_delivery": td}
    return data




# ─── Cell writer ─────────────────────────────────────────────────
def _w(ws, row, col, value, font=None, fill=None, num_fmt=None, na=False):
    c = ws.cell(row=row, column=col)
    if na:
        c.value = "N/A"
        c.font = FONT_NA
    else:
        c.value = value
        c.font = font or FONT_DATA
    if fill: c.fill = fill
    c.alignment = ALIGN_CENTER
    if num_fmt: c.number_format = num_fmt
    return c


def _header(ws, title, months, sc=2):
    n = len(months)
    ec = min(sc + 11, sc + n - 1)
    ws.cell(row=1, column=sc, value=title).font = FONT_TITLE
    ws.cell(row=1, column=sc).fill = FILL_ALT
    ws.cell(row=1, column=sc).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
    for c in range(sc, ec + 1): ws.cell(row=1, column=c).fill = FILL_ALT

    sections = [(sc, min(sc+2, sc+n-1), "PAST 90 DAYS"),
                (sc+3, min(sc+5, sc+n-1), "PAST 6 MONTHS"),
                (sc+6, min(sc+11, sc+n-1), "PAST 12 MONTHS")]
    for s, e, label in sections:
        if s <= sc + n - 1:
            ws.cell(row=2, column=s, value=label).font = FONT_SECTION_LABEL
            ws.cell(row=2, column=s).fill = FILL_ALT
            ws.cell(row=2, column=s).alignment = ALIGN_CENTER
            if e > s and e <= sc + n - 1:
                ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
                for c in range(s, e+1): ws.cell(row=2, column=c).fill = FILL_ALT

    for i, m in enumerate(months):
        c = ws.cell(row=3, column=sc+i, value=month_key_to_label(m))
        c.font = FONT_MONTH
        c.alignment = ALIGN_CENTER
        c.number_format = '@'

    ws.column_dimensions['A'].width = 16
    if sc == 3: ws.column_dimensions['B'].width = 4
    for i in range(n): ws.column_dimensions[get_column_letter(sc+i)].width = 13


def _label(ws, row, label, bold=True):
    ws.cell(row=row, column=1, value=label).font = FONT_ROW_LABEL if bold else FONT_SUB_LABEL
    ws.cell(row=row, column=1).fill = FILL_ALT
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER


# ─── Sheet Builders ──────────────────────────────────────────────
def build_views(wb, months, yt_ct, mega, ep_counts={}):
    ws = wb.create_sheet("📺 Views")
    ws.sheet_properties.tabColor = "00FF00"
    n = len(months)
    _header(ws, "📺 VIEWS GROWTH", months)

    rows = [(4, "# OF EPS", True), (5, "YT VIDS", False), (6, "YT SHORTS", False),
            (7, "YT LIVES", False), (8, "SPOTIFY", False), (9, "TOTAL", True)]

    for r, label, is_total_row in rows:
        _label(ws, r, label)
        for i, m in enumerate(months):
            col = i + 2
            ct = yt_ct.get(m, {})
            alt = r in (4, 9)
            fill = FILL_ALT if alt else FILL_NONE

            if label == "# OF EPS":
                n_eps = ep_counts.get(m)
                if n_eps:
                    _w(ws, r, col, n_eps, fill=fill, num_fmt=NUM_FMT)
                else:
                    _w(ws, r, col, None, fill=fill, na=True)
            elif label == "YT VIDS":
                v = ct.get("VIDEO_ON_DEMAND")
                _w(ws, r, col, v, fill=fill, num_fmt=NUM_FMT) if v else _w(ws, r, col, None, fill=fill, na=True)
            elif label == "YT SHORTS":
                v = ct.get("SHORTS")
                _w(ws, r, col, v, fill=fill, num_fmt=NUM_FMT) if v else _w(ws, r, col, None, fill=fill, na=True)
            elif label == "YT LIVES":
                v = ct.get("LIVE_STREAM")
                _w(ws, r, col, v if v else 0, fill=fill, num_fmt=NUM_FMT) if m in yt_ct else _w(ws, r, col, None, fill=fill, na=True)
            elif label == "SPOTIFY":
                md = mega.get(m, {})
                st = md.get("streams", 0) if isinstance(md, dict) else 0
                _w(ws, r, col, st, fill=fill, num_fmt=NUM_FMT) if st > 0 else _w(ws, r, col, None, fill=fill, na=True)
            elif label == "TOTAL":
                cl = get_column_letter(col)
                _w(ws, r, col, f"=SUM({cl}5:{cl}8)", fill=fill, num_fmt=NUM_FMT)


def build_listens(wb, months, mega, ep_counts={}):
    ws = wb.create_sheet("🎧 Listens")
    ws.sheet_properties.tabColor = "00FF00"
    _header(ws, "🎧 LISTENING GROWTH", months)

    for r, label in [(4, "# OF EPS"), (5, "PER EP LISTENS"), (6, "TOTAL LISTENS")]:
        _label(ws, r, label)
        for i, m in enumerate(months):
            col = i + 2
            md = mega.get(m, {})
            dl = md.get("downloads", 0) if isinstance(md, dict) else 0
            n_eps = ep_counts.get(m)
            if label == "# OF EPS":
                _w(ws, r, col, n_eps, fill=FILL_ALT, num_fmt=NUM_FMT) if n_eps else _w(ws, r, col, None, fill=FILL_ALT, na=True)
            elif label == "PER EP LISTENS":
                per_ep = round(dl / n_eps) if (dl > 0 and n_eps) else None
                _w(ws, r, col, per_ep, fill=FILL_ALT, num_fmt=NUM_FMT) if per_ep else _w(ws, r, col, None, fill=FILL_ALT, na=True)
            elif label == "TOTAL LISTENS" and dl > 0:
                _w(ws, r, col, dl, fill=FILL_ALT, num_fmt=NUM_FMT)
            else:
                _w(ws, r, col, None, fill=FILL_ALT, na=True)


def build_socials(wb, months):
    ws = wb.create_sheet("📱Socials")
    ws.sheet_properties.tabColor = "00FF00"
    _header(ws, "📱 SOCIAL GROWTH", months)
    n = len(months)

    structure = [
        (4, "VIEWS PER POST AVG", True), (5, "TOTAL FOLLOWERS", True),
        (6, "TWITTER/X", False), (7, "INSTAGRAM", False), (8, "TIKTOK", False),
        (9, "TOTAL", True), (10, "FOLLOWER GAIN (NET)", True),
        (11, "TWITTER/X", False), (12, "INSTAGRAM", False), (13, "TIKTOK", False),
        (14, "TOTAL", True), (15, "TOTAL POSTS ", True),
        (16, "TWITTER/X", False), (17, "INSTAGRAM", False), (18, "TIKTOK", False),
        (19, "TOTAL", True), (20, "TOTAL VIEWS", True),
        (21, "TWITTER/X", False), (22, "INSTAGRAM", False), (23, "TIKTOK", False),
        (24, "TOTAL", True), (25, "TOP POSTS", True),
        (26, "TWITTER/X", False), (27, "INSTAGRAM", False), (28, "TIKTOK", False),
        (29, "ENGAGEMENT RATE", True), (30, "INSTAGRAM", False),
        (31, "TWITTER/X", False), (32, "TIKTOK", False), (33, "TOTAL", True),
    ]
    for r, label, bold in structure:
        _label(ws, r, label, bold)
        for i in range(n):
            _w(ws, r, i+2, None, fill=FILL_ALT if bold else FILL_NONE, na=True)


def build_subs(wb, months, yt_monthly, current_subs):
    ws = wb.create_sheet("👥 Subs")
    ws.sheet_properties.tabColor = "00FF00"
    _header(ws, "👥 SUBSCRIBER GROWTH", months)
    n = len(months)

    # Reconstruct historical sub totals by working backwards
    sub_totals = {}
    running = current_subs
    for m in months:
        sub_totals[m] = running
        if m in yt_monthly:
            running -= (yt_monthly[m]["subs_gained"] - yt_monthly[m]["subs_lost"])

    structure = [
        (4, "YOUTUBE", True, None), (5, "TOTAL", True, "yt_total"),
        (6, "(+/-)", False, "yt_delta"), (7, "AUDIO", True, None),
        (8, "APPLE", False, "na"), (9, "SPOTIFY", False, "na"),
        (10, "TOTAL", True, "audio_sum"), (11, "(+/-)", False, "audio_delta"),
        (12, "TOTAL", True, "combined"),
    ]

    for r, label, bold, key in structure:
        _label(ws, r, label, bold)
        for i, m in enumerate(months):
            col = i + 2
            cl = get_column_letter(col)
            ncl = get_column_letter(col + 1) if i < n - 1 else None
            fill = FILL_ALT if bold else FILL_NONE

            if key == "yt_total":
                _w(ws, r, col, sub_totals.get(m), fill=fill, num_fmt=NUM_FMT) if m in sub_totals else _w(ws, r, col, None, fill=fill, na=True)
            elif key == "yt_delta":
                _w(ws, r, col, f"={cl}5-{ncl}5", fill=fill, num_fmt=NUM_FMT) if ncl else _w(ws, r, col, None, fill=fill, na=True)
            elif key == "na":
                _w(ws, r, col, None, fill=fill, na=True)
            elif key == "audio_sum":
                _w(ws, r, col, f"=SUM({cl}8:{cl}9)", fill=fill, num_fmt=NUM_FMT)
            elif key == "audio_delta":
                _w(ws, r, col, f"={cl}10-{ncl}10", fill=fill, num_fmt=NUM_FMT) if ncl else _w(ws, r, col, None, fill=fill, na=True)
            elif key == "combined":
                _w(ws, r, col, f"=SUM({cl}5,{cl}10)", fill=fill, num_fmt=NUM_FMT)


def build_kpis(wb, months, yt_ct, shorts_count={}, reporting_data={}):
    ws = wb.create_sheet("🔑 KPIs")
    ws.sheet_properties.tabColor = "00FF00"
    _header(ws, "🔑  KPIs (YOUTUBE)", months)

    structure = [
        (4, "CTR", True, None), (5, "VIDEOS", False, "manual_pct"),
        (6, "SHORTS", False, "manual_pct"), (7, "SHORTS", True, None),
        (8, "# OF SHORTS", False, "shorts_count"), (9, "% OF VIEWS", False, "shorts_pct"),
        (10, "SEARCH", True, None), (11, "% OF VIEWS", False, "manual_pct"),
        (12, "CTR IN SEARCH", False, "manual_pct"), (13, "TV VIEWS", True, None),
        (14, "VIDEOS", False, "manual_pct"), (15, "LIVES", False, "manual_pct"),
        (16, "TOTAL", False, "manual_pct"),
    ]

    for r, label, bold, key in structure:
        _label(ws, r, label, bold)
        for i, m in enumerate(months):
            col = i + 2
            ct = yt_ct.get(m, {})
            fill = FILL_ALT if bold else FILL_NONE

            rep = reporting_data.get(m, {}) if m in reporting_data else {}
            if key in ("ctr_video", "ctr_shorts"):
                v = rep.get("ctr")
                if v:
                    _w(ws, r, col, v, fill=fill, num_fmt=PCT_FMT)
                else:
                    c = _w(ws, r, col, None, fill=fill, na=True)
                    c.number_format = PCT_FMT
            elif key == "shorts_pct":
                s = ct.get("SHORTS", 0)
                t = ct.get("VIDEO_ON_DEMAND", 0) + s + ct.get("LIVE_STREAM", 0)
                if t > 0 and m in yt_ct:
                    _w(ws, r, col, round(s / t, 3), fill=fill, num_fmt=PCT_FMT)
                else:
                    _w(ws, r, col, None, fill=fill, na=True)
            elif key == "manual_pct":
                c = _w(ws, r, col, None, fill=fill, na=True)
                c.number_format = PCT_FMT
            elif key == "shorts_count":
                n = shorts_count.get(m)
                _w(ws, r, col, n, fill=fill, num_fmt=NUM_FMT) if n else _w(ws, r, col, None, fill=fill, na=True)
            elif key == "manual":
                _w(ws, r, col, None, fill=fill, na=True)


def build_charts(wb, months):
    ws = wb.create_sheet("🏆 Charts")
    ws.sheet_properties.tabColor = "00FF00"
    _header(ws, "🏆 CHARTS", months)
    for r, label in [(4, "APPLE"), (5, "SPOTIFY"), (6, "YOUTUBE")]:
        _label(ws, r, label)
        for i in range(len(months)):
            _w(ws, r, i+2, None, fill=FILL_ALT, na=True)


def build_episodics(wb, months, yt_ct):
    ws = wb.create_sheet("▶️ Episodics")
    ws.sheet_properties.tabColor = "FF9900"
    n = len(months)
    _header(ws, "📈 EPISODIC GROWTH", months, sc=3)

    # Section header: VIEWS PER / TAPED + LIVE VIEWS
    ws.cell(row=4, column=1, value="VIEWS PER").font = FONT_ROW_LABEL
    ws.cell(row=4, column=1).fill = FILL_SECTION
    ec = min(3 + n - 1, 14)
    ws.cell(row=4, column=3, value="TAPED + LIVE VIEWS").font = FONT_ROW_LABEL
    ws.cell(row=4, column=3).fill = FILL_SECTION
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=ec)
    for c in range(3, ec + 1): ws.cell(row=4, column=c).fill = FILL_SECTION

    for cr, ar, label in [(5, 6, "LONG-FORM"), (7, 8, "MID-FORM"), (9, 10, "SHORT-FORM")]:
        _label(ws, cr, label, bold=False)
        for i in range(n):
            _w(ws, cr, i+3, None, fill=FILL_ALT, na=True)  # Count row
            _w(ws, ar, i+3, None, na=True)  # Avg views row

    _label(ws, 11, "TOTAL PACKAGE", bold=False)
    for i, m in enumerate(months):
        ct = yt_ct.get(m, {})
        t = ct.get("VIDEO_ON_DEMAND", 0) + ct.get("SHORTS", 0) + ct.get("LIVE_STREAM", 0)
        if t > 0:
            _w(ws, 11, i+3, t, fill=FILL_ALT, num_fmt=NUM_FMT)
        else:
            _w(ws, 11, i+3, None, fill=FILL_ALT, na=True)

    # Lives section
    ws.cell(row=12, column=1, value="LIVES").font = FONT_ROW_LABEL
    ws.cell(row=12, column=1).fill = FILL_SECTION
    ws.cell(row=12, column=3, value="LIVE VIEWS").font = FONT_ROW_LABEL
    ws.cell(row=12, column=3).fill = FILL_SECTION
    ws.merge_cells(start_row=12, start_column=3, end_row=12, end_column=ec)
    for c in range(3, ec + 1): ws.cell(row=12, column=c).fill = FILL_SECTION

    for r, label in [(13, "PER EP"), (14, "TOTAL"), (15, "# OF EPS")]:
        _label(ws, r, label, bold=False)
        for i, m in enumerate(months):
            ct = yt_ct.get(m, {})
            if label == "TOTAL" and m in yt_ct:
                v = ct.get("LIVE_STREAM", 0)
                _w(ws, r, i+3, v, fill=FILL_ALT, num_fmt=NUM_FMT) if v else _w(ws, r, i+3, None, fill=FILL_ALT, na=True)
            else:
                _w(ws, r, i+3, None, fill=FILL_ALT, na=True)


# ─── Main ────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Road Trippin' Analytics Tracker Builder v2")
    parser.add_argument("--megaphone-monthly",  default=None, help="Override default path to megaphone_monthly.csv")
    parser.add_argument("--megaphone-episodes", default=None, help="Override default path to megaphone_per_episode.csv")
    parser.add_argument("--output", default="Road_Trippin_Analytics_Tracker.xlsx")
    parser.add_argument("--months", type=int, default=22)
    args = parser.parse_args()

    # Always use absolute paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "data")

    # Fixed filenames — just rename your Megaphone exports to these each month
    if not args.megaphone_monthly:
        args.megaphone_monthly = os.path.join(data_dir, "megaphone_monthly.csv")
    if not args.megaphone_episodes:
        args.megaphone_episodes = os.path.join(data_dir, "megaphone_per_episode.csv")

    print("=" * 60)
    print("ROAD TRIPPIN' — ANALYTICS TRACKER BUILDER v2")
    print("=" * 60)

    months = get_month_keys(args.months)
    print(f"Months: {month_key_to_label(months[0])} ({months[0]}) → {month_key_to_label(months[-1])} ({months[-1]})")
    print(f"Total: {len(months)} months\n")

    print("[1/3] YouTube Analytics API...")
    yt_monthly, yt_content, current_subs, yt_shorts_count, reporting_data, best_of_data, date_range_label = {}, {}, 0, {}, {}, {}, datetime.now().strftime('%b %Y')
    top_content_monthly = {}
    try:
        creds = authenticate_yt()
        youtube = build_api("youtube", "v3", credentials=creds)
        yt_analytics = build_api("youtubeAnalytics", "v2", credentials=creds)
        start = f"{months[-1]}-01"
        end = datetime.now().replace(day=1).strftime("%Y-%m-%d")
        print(f"  Date range: {start} → {end}")
        print("  Pulling monthly totals...")
    except Exception as e:
        print(f"  ✗ Auth failed: {e}")
        creds = None

    if creds:
        try:
            yt_monthly = pull_yt_monthly(yt_analytics, start, end)
            print(f"  ✓ Monthly totals: {len(yt_monthly)} months")
        except Exception as e:
            print(f"  ✗ Monthly totals failed: {e}")

        try:
            yt_content = pull_yt_by_content_type(yt_analytics, start, end)
            print(f"  ✓ Content types: {len(yt_content)} months")
        except Exception as e:
            print(f"  ✗ Content types failed: {e}")

        try:
            current_subs = pull_yt_subscriber_total(youtube)
            print(f"  ✓ Subscribers: {current_subs:,}")
        except Exception as e:
            print(f"  ✗ Subscribers failed: {e}")

        try:
            print("  Counting Shorts by month (Data API)...")
            yt_shorts_count = pull_yt_shorts_count(youtube)
            print(f"  ✓ Shorts count: {len(yt_shorts_count)} months")
        except Exception as e:
            print(f"  ✗ Shorts count failed: {e}")
            yt_shorts_count = {}

        try:
            print("  Pulling Best Of (Data API)...")
            from datetime import datetime as _dt
            now = _dt.now()
            start_mo = (now.replace(day=1) - relativedelta(months=0)).strftime("%b %d").lstrip("0")
            end_mo = now.strftime("%b %d, %Y")
            date_range_label = f"{start_mo} – {end_mo}"
            best_of_data = pull_best_of(youtube, months, n_months=1)
            total_best = sum(len(v) for v in best_of_data.values())
            print(f"  ✓ Best Of: {total_best} videos across 3 categories")
        except Exception as e:
            print(f"  ✗ Best Of failed: {e}")
            best_of_data = {}
            date_range_label = datetime.now().strftime("%b %Y")

        try:
            print("  Pulling Top Content Monthly (filterable history)...")
            top_content_monthly = pull_top_content_monthly(youtube, months, top_n=10)
            total_tc = sum(sum(len(b) for b in v.values()) for v in top_content_monthly.values())
            print(f"  ✓ Top Content: {total_tc} videos across {len(top_content_monthly)} months")
        except Exception as e:
            print(f"  ✗ Top Content Monthly failed: {e}")
            top_content_monthly = {}

        try:
            print("  Pulling Reporting API (CTR, traffic, device)...")
            reporting_data = pull_reporting_api(creds, data_dir)
            has_ctr = any(v.get("ctr") for v in reporting_data.values())
            print(f"  ✓ Reporting: {len(reporting_data)} months | CTR available: {has_ctr}")
        except Exception as e:
            print(f"  ✗ Reporting API failed: {e}")
            reporting_data = {}

    print("\n[2/3] Megaphone...")
    print(f"  Monthly CSV: {args.megaphone_monthly}")
    mega = load_megaphone_monthly(args.megaphone_monthly)
    print(f"  ✓ {len(mega)} months: {sorted(mega.keys())}" if mega else "  ⚠ Monthly CSV not found — rename your export to megaphone_monthly.csv")

    # Episode counts come from the Megaphone API (pulled in [1/3] via pull_megaphone_api)
    # Per-episode CSV is no longer needed
    mega_ep_counts = {}
    if MEGAPHONE_TOKEN:
        try:
            _, mega_ep_counts = pull_megaphone_api(
                MEGAPHONE_TOKEN, MEGAPHONE_NETWORK_ID, MEGAPHONE_PODCAST_ID,
                n_months=args.months
            )
            print(f"  ✓ Episode counts from API: {sum(mega_ep_counts.values())} total episodes")
        except Exception as e:
            print(f"  ⚠ Episode count API failed: {e}")
    if not mega_ep_counts:
        print(f"  ⚠ No episode counts available — # OF EPS will be N/A")



    print(f"\nBuilding tracker...")
    wb = Workbook()
    wb.remove(wb.active)

    build_views(wb, months, yt_content, mega, mega_ep_counts)
    build_listens(wb, months, mega, mega_ep_counts)
    build_socials(wb, months)
    build_subs(wb, months, yt_monthly, current_subs)
    build_kpis(wb, months, yt_content, yt_shorts_count, reporting_data)
    build_charts(wb, months)
    build_episodics(wb, months, yt_content)
    build_best_of(wb, best_of_data, date_range_label)

    out = os.path.join(data_dir, args.output) if os.path.isdir(data_dir) else os.path.join(script_dir, args.output)
    wb.save(out)
    print(f"\n✓ Saved: {out}")

    # Also write tracker_data.json — will replace xlsx entirely in a future update
    def _clean(v):
        if v is None: return None
        if isinstance(v, (int, float)): return v
        s = str(v).strip()
        if s in ('N/A', '—', '', 'None'): return None
        try: return float(s.replace(',', '')) if '.' in s else int(s.replace(',', ''))
        except: return s

    def _series(src, key, ms):
        return {m: _clean(src.get(m, {}).get(key)) for m in ms}

    tracker_json = {
        "generated": datetime.now().isoformat(),
        "months": months,
        "yt": {
            # yt_content keyed by month → {VIDEO_ON_DEMAND, SHORTS, LIVE_STREAM}
            "vids":        {m: _clean(yt_content.get(m, {}).get("VIDEO_ON_DEMAND")) for m in months},
            "shorts":      {m: _clean(yt_content.get(m, {}).get("SHORTS"))          for m in months},
            "lives":       {m: _clean(yt_content.get(m, {}).get("LIVE_STREAM"))     for m in months},
            # yt_monthly keyed by month → {views, subs_gained, subs_lost, ...}
            "subs_gained": {m: _clean(yt_monthly.get(m, {}).get("subs_gained"))     for m in months},
            "subs_lost":   {m: _clean(yt_monthly.get(m, {}).get("subs_lost"))       for m in months},
        },
        "audio": {
            "downloads": {m: _clean(mega.get(m, {}).get("downloads")) for m in months},
            "streams":   {m: _clean(mega.get(m, {}).get("streams")) for m in months},
            "episodes":  {m: mega_ep_counts.get(m) for m in months},
        },
        "kpis": {
            "shorts_count": {m: yt_shorts_count.get(m) for m in months},
            "ctr":          {m: reporting_data.get(m, {}).get("ctr") for m in months},
            "search_pct":   {m: reporting_data.get(m, {}).get("search_pct") for m in months},
            "tv_pct":       {m: reporting_data.get(m, {}).get("tv_pct") for m in months},
        },
        "best_of": {
            "label": date_range_label,
            "long":  best_of_data.get("long",  []),
            "mid":   best_of_data.get("mid",   []),
            "short": best_of_data.get("short", []),
        },
        "top_content_monthly": top_content_monthly,
        "current_subs": current_subs,
    }

    import json as _json
    json_out = os.path.join(data_dir, "tracker_data.json")
    with open(json_out, "w", encoding="utf-8") as f:
        _json.dump(tracker_json, f, indent=2, default=str)
    print(f"✓ JSON:  {json_out}")
    print("=" * 60)


if __name__ == "__main__":
    main()